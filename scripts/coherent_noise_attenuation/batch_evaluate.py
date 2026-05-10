"""Batch evaluation: iterate experiment directories, run inference on held-out
test sets, compute metrics before/after denoising, and output an Excel workbook
with one sheet per noise level.

Usage::

    python scripts/coherent_noise_attenuation/batch_evaluate.py \\
        --root_dir results/coherent_noise_attenuation \\
        --output results/coherent_noise_attenuation/batch_evaluation.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import torch

# ---------------------------------------------------------------------------
# repo-root bootstrap (same pattern as the training scripts)
# ---------------------------------------------------------------------------
_REPO_ROOT = next(
    (
        p
        for p in Path(__file__).resolve().parents
        if (p / "model").is_dir() and (p / "utils").is_dir()
    ),
    None,
)
if _REPO_ROOT is None:
    raise RuntimeError(
        "Cannot find repo root (a directory containing both ``model/`` and ``utils/``)."
    )
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from model.coherent_noise_attenuation import build_model  # noqa: E402
from utils.inference_utils import compute_shot_metrics, inference_on_shots  # noqa: E402

# ---------------------------------------------------------------------------
# constants
# ---------------------------------------------------------------------------
METRIC_NAMES = ["snr", "psnr", "ssim", "mae", "mse", "rmse"]
METRIC_DISPLAY = [m.upper() for m in METRIC_NAMES]  # SNR, PSNR, SSIM, MAE, MSE, RMSE

# registry key → display name (rows, top-to-bottom in each sheet)
MODEL_DISPLAY = {
    "unet": "UNet",
    "res_unet": "ResUNet",
    "dncnn": "DnCNN",
    "atten_unet": "Attention UNet",
}
MODEL_ROW_ORDER = ["unet", "res_unet", "dncnn", "atten_unet"]

# directory-name pattern: denoise_{model}_base{date}_level{level}_seed{seed}
_DIR_RE = re.compile(
    r"^denoise_(.+)_base\d+_level([\d.]+)_seed(\d+)$"
)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def parse_dir_name(name: str) -> Optional[Tuple[str, str, str]]:
    """Return ``(model_name, noise_level, seed)`` or ``None`` if unparseable."""
    m = _DIR_RE.match(name)
    if m is None:
        return None
    return m.group(1), m.group(2), m.group(3)


def discover_results(root: Path) -> List[Dict[str, Any]]:
    """Scan *root* for experiment directories and return metadata dicts.

    Each dict contains: ``dir`` (Path), ``model``, ``level``, ``seed``.
    Only directories with both ``checkpoints/best.pt`` and ``test_set/``
    are kept.
    """
    entries: List[Dict[str, Any]] = []
    for d in sorted(root.iterdir()):
        if not d.is_dir():
            continue
        parsed = parse_dir_name(d.name)
        if parsed is None:
            print(f"[SKIP] cannot parse directory name: {d.name}")
            continue
        ckpt = d / "checkpoints" / "best.pt"
        test_dir = d / "test_set"
        if not ckpt.is_file():
            print(f"[SKIP] missing best.pt: {d.name}")
            continue
        if not test_dir.is_dir():
            print(f"[SKIP] missing test_set/: {d.name}")
            continue
        entries.append(
            {
                "dir": d,
                "model": parsed[0],
                "level": parsed[1],
                "seed": parsed[2],
            }
        )
    return entries


def load_model_from_checkpoint(ckpt_path: Path, device: torch.device) -> torch.nn.Module:
    """Load a model from a training checkpoint."""
    ckpt = torch.load(ckpt_path, map_location=device, weights_only=False)
    model_cfg = ckpt["extras"]["config"]["model"]
    model = build_model(model_cfg)
    model.load_state_dict(ckpt["model"])
    model.to(device)
    model.eval()
    return model


def evaluate_one(
    result_dir: Path,
    device: torch.device,
    patch_size: Tuple[int, int] = (128, 256),
    overlap: float = 0.5,
    batch_size: int = 8,
) -> Optional[Dict[str, Any]]:
    """Run inference + metric computation for a single experiment directory.

    Returns a dict of before/after metrics or ``None`` on failure.
    """
    ckpt_path = result_dir / "checkpoints" / "best.pt"
    test_dir = result_dir / "test_set"

    # --- load model ---------------------------------------------------------
    try:
        model = load_model_from_checkpoint(ckpt_path, device)
        total_params = sum(p.numel() for p in model.parameters())
        num_params_m = total_params / 1e6
    except Exception:
        print(f"[ERROR] failed to load model from {ckpt_path}:")
        traceback.print_exc()
        return None

    # --- load test data -----------------------------------------------------
    try:
        input_shots = np.load(test_dir / "input_shots.npy")
        target_shots = np.load(test_dir / "target_shots.npy")
    except Exception:
        print(f"[ERROR] failed to load test data from {test_dir}:")
        traceback.print_exc()
        return None

    # --- inference ----------------------------------------------------------
    try:
        pred_noise = inference_on_shots(
            model,
            input_shots,
            patch_size=patch_size,
            overlap=overlap,
            device=device,
            batch_size=batch_size,
        )
    except Exception:
        print(f"[ERROR] inference failed for {result_dir.name}:")
        traceback.print_exc()
        return None

    # --- compute signals (3D) ------------------------------------------------
    clean_ref = input_shots - target_shots          # ground-truth clean
    denoised = input_shots - pred_noise             # model output

    # --- flatten to 2D: (n_shots × n_traces, n_time) ------------------------
    n_time = input_shots.shape[-1]
    noisy_2d = input_shots.reshape(-1, n_time)
    clean_2d = clean_ref.reshape(-1, n_time)
    denoised_2d = denoised.reshape(-1, n_time)

    # --- metrics (reshape 2D → 3D for compute_shot_metrics) -----------------
    _, before_mean = compute_shot_metrics(
        noisy_2d.reshape(1, -1, n_time), clean_2d.reshape(1, -1, n_time),
        METRIC_NAMES,
    )
    _, after_mean = compute_shot_metrics(
        denoised_2d.reshape(1, -1, n_time), clean_2d.reshape(1, -1, n_time),
        METRIC_NAMES,
    )

    return {
        "before": before_mean,
        "after": after_mean,
        "num_params_m": num_params_m,
    }


def aggregate(
    entries: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """Aggregate per-seed results into mean±std per (level, model).

    Returns ``{level: {model: {metric: (mean, std)}, ..., "raw": {metric: val}}}``.
    """
    # group: level → model → list of per-seed metric dicts
    groups: Dict[str, Dict[str, List[Dict[str, float]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    raw_by_level: Dict[str, Dict[str, float]] = {}
    params_by_model: Dict[str, float] = {}

    for entry in entries:
        level = entry["level"]
        model = entry["model"]
        after = entry.get("after", {})
        if after:
            groups[level][model].append(after)
        # capture raw (before) metrics — deterministic, just take the first
        before = entry.get("before", {})
        if before and level not in raw_by_level:
            raw_by_level[level] = before
        # param count — same for all seeds of a model, take first
        n_params = entry.get("num_params_m")
        if n_params is not None and model not in params_by_model:
            params_by_model[model] = n_params

    # compute mean±std
    aggregated: Dict[str, Dict[str, Any]] = {}
    for level in sorted(groups.keys(), key=float):
        aggregated[level] = {}
        # raw row
        if level in raw_by_level:
            aggregated[level]["raw"] = raw_by_level[level]
        for model in MODEL_ROW_ORDER:
            seeds = groups[level].get(model, [])
            if not seeds:
                continue
            model_stats: Dict[str, Tuple[float, float]] = {}
            for m in METRIC_NAMES:
                vals = [s[m] for s in seeds if m in s]
                if len(vals) == 0:
                    continue
                mean = float(np.mean(vals))
                if len(vals) >= 2:
                    std = float(np.std(vals, ddof=1))
                else:
                    std = 0.0
                model_stats[m] = (mean, std)
            aggregated[level][model] = model_stats

    return aggregated, params_by_model


def _fmt(mean: float, std: float, metric: str) -> str:
    """Format a cell value as ``mean±std`` with metric-appropriate precision.

    MAE, MSE, RMSE: mean to 6 decimal places, std to 2 decimal places.
    SNR, PSNR, SSIM: mean and std to 2 decimal places.
    """
    if metric in ("mae", "mse", "rmse"):
        return f"{mean:.6f}±{std:.2f}"
    return f"{mean:.2f}±{std:.2f}"


def build_excel(
    aggregated: Dict[str, Dict[str, Any]],
    params_by_model: Dict[str, float],
    output_path: Path,
) -> None:
    """Write one sheet per noise level. Rows = methods, columns = metrics."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError:
        print(
            "openpyxl is required for Excel output. Install it with:\n"
            "    pip install openpyxl"
        )
        raise

    wb = openpyxl.Workbook()
    # remove default sheet — we'll create one per level
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    n_metric_cols = len(METRIC_DISPLAY)

    for level in sorted(aggregated.keys(), key=float):
        ws = wb.create_sheet(title=f"Noise {level}")

        # column A = Method, B = Parameters (M), C..H = metrics
        ws.cell(row=1, column=1, value="Method").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).border = thin_border

        cell = ws.cell(row=1, column=2, value="Parameters (M)")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        for ci, m_disp in enumerate(METRIC_DISPLAY, start=3):
            cell = ws.cell(row=1, column=ci, value=m_disp)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        level_data = aggregated[level]

        # row order: raw, then each model in MODEL_ROW_ORDER
        row_order = ["raw"] + [m for m in MODEL_ROW_ORDER if m in level_data]
        for ri, key in enumerate(row_order, start=2):
            # method label
            label = "Raw (noisy)" if key == "raw" else MODEL_DISPLAY.get(key, key)
            cell = ws.cell(row=ri, column=1, value=label)
            cell.font = Font(bold=True) if key == "raw" else Font()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

            # param count
            if key == "raw":
                cell = ws.cell(row=ri, column=2, value="—")
            else:
                n_m = params_by_model.get(key)
                cell = ws.cell(row=ri, column=2,
                               value=round(n_m, 2) if n_m is not None else "—")
            cell.alignment = center_align
            cell.border = thin_border

            data = level_data[key]
            for ci, m_name in enumerate(METRIC_NAMES, start=3):
                if m_name not in data:
                    cell = ws.cell(row=ri, column=ci, value="—")
                elif key == "raw":
                    val = data[m_name]
                    cell = ws.cell(row=ri, column=ci, value=_fmt(val, 0.0, m_name))
                else:
                    mean, std = data[m_name]
                    cell = ws.cell(row=ri, column=ci, value=_fmt(mean, std, m_name))
                cell.alignment = center_align
                cell.border = thin_border

        # auto-width
        from openpyxl.utils import get_column_letter
        for ci in range(1, n_metric_cols + 3):
            max_w = 0
            for row in ws.iter_rows(min_col=ci, max_col=ci):
                for c in row:
                    if c.value is not None:
                        max_w = max(max_w, len(str(c.value)))
            ws.column_dimensions[get_column_letter(ci)].width = max_w + 4

        ws.freeze_panes = "C2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved Excel to: {output_path}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Batch evaluate denoising results on held-out test sets"
    )
    parser.add_argument(
        "--root_dir", default="/data/shared/benchmark/ground_roll/results", type=Path,
        help="Directory containing experiment subdirectories",
    )
    parser.add_argument(
        "--output", default="/data/shared/benchmark/ground_roll/results/batch_evaluation.xlsx", type=Path,
        help="Path for the output Excel file (.xlsx)",
    )
    parser.add_argument(
        "--device", default="cuda:7" if torch.cuda.is_available() else "cpu",
        help="Device for inference (default: cuda if available, else cpu)",
    )
    parser.add_argument(
        "--batch_size", type=int, default=8,
        help="Batch size for inference (default: 8)",
    )
    args = parser.parse_args()

    root_dir: Path = args.root_dir
    if not root_dir.is_dir():
        print(f"Root directory does not exist: {root_dir}")
        sys.exit(1)

    device = torch.device(args.device)
    print(f"Using device: {device}")

    # --- discover -----------------------------------------------------------
    entries = discover_results(root_dir)
    if not entries:
        print("No valid experiment directories found.")
        sys.exit(0)

    print(f"Found {len(entries)} experiment(s) to evaluate.")

    # --- evaluate -----------------------------------------------------------
    for i, entry in enumerate(entries):
        d = entry["dir"]
        print(
            f"[{i + 1}/{len(entries)}] {d.name}  "
            f"(model={entry['model']}, level={entry['level']}, seed={entry['seed']})"
        )
        result = evaluate_one(
            d,
            device=device,
            batch_size=args.batch_size,
        )
        if result is None:
            print(f"  -> FAILED, skipping")
            entry["before"] = {}
            entry["after"] = {}
        else:
            entry["before"] = result["before"]
            entry["after"] = result["after"]
            entry["num_params_m"] = result["num_params_m"]
            b = result["before"]
            a = result["after"]
            print(
                f"  Params: {result['num_params_m']:.2f}M  |  "
                f"SNR:  {b['snr']:>7.2f} -> {a['snr']:>7.2f} dB  |  "
                f"PSNR: {b['psnr']:>7.2f} -> {a['psnr']:>7.2f} dB  |  "
                f"SSIM: {b['ssim']:.4f} -> {a['ssim']:.4f}  |  "
                f"MSE:  {b['mse']:.6f} -> {a['mse']:.6f}"
            )

    # --- aggregate per level -------------------------------------------------
    aggregated, params_by_model = aggregate(entries)

    # summary of what was aggregated
    print("Model parameter counts:")
    for m in MODEL_ROW_ORDER:
        if m in params_by_model:
            print(f"  {MODEL_DISPLAY.get(m, m):<16s}  {params_by_model[m]:.2f}M")
    for level in sorted(aggregated.keys(), key=float):
        models = [k for k in aggregated[level] if k != "raw"]
        print(f"Noise {level}: raw + {len(models)} model(s) — {', '.join(MODEL_DISPLAY.get(m, m) for m in models)}")

    # --- export -------------------------------------------------------------
    build_excel(aggregated, params_by_model, args.output)


if __name__ == "__main__":
    main()

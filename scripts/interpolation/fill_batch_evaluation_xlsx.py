#!/usr/bin/env python
"""Fill interpolation benchmark results into an Excel workbook.

For every experiment config under collected/configs/ (one YAML per
experiment/seed), the per-seed inference outputs produced by
run_batch_best_inference.sh (collected/<name>/inference/metrics_summary.json)
are read, grouped by setting (seed stripped from the name), and aggregated as
mean±std across seeds.  One row per setting is written into an "Interpolation"
sheet of the target workbook, preserving any existing sheets (e.g. "Multiples").

Metrics in metrics_summary.json are mapped to workbook columns by uppercasing
the JSON key (``eb_wse_very_weak_5_20_ne`` -> ``EB_WSE_VERY_WEAK_5_20_NE``).
RMSE is derived as sqrt(MSE) per seed.  Frequency-range metadata columns are
left blank (—), matching the existing benchmark rows.

Parameters (M) is counted from the checkpoint state_dict (one load per model
type, cached).

Usage::

    python scripts/interpolation/fill_batch_evaluation_xlsx.py
    python scripts/interpolation/fill_batch_evaluation_xlsx.py \
        --config-dir collected/configs \
        --collect-root collected \
        --params-dir collected/params \
        --output batch_evaluation_part.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import yaml

_REPO_ROOT = Path(__file__).resolve().parents[2]

_NAME_RE = re.compile(r"^interp_(.+)_seed([0-9]+)_(.+)$")

# Fallback column headers used when the target workbook has no existing sheet
# to copy from.  Must mirror the "Multiples" sheet layout.
HEADERS = [
    "Method", "Parameters (M)",
    "SNR", "PSNR", "SSIM", "MAE", "MSE", "RMSE",
    "EB_WSE_MEDIUM_40_70_NE", "EB_WSE_MEDIUM_40_70_SNR",
    "EB_WSE_STRONG_70_100_NE", "EB_WSE_STRONG_70_100_SNR",
    "EB_WSE_VERY_WEAK_5_20_NE", "EB_WSE_VERY_WEAK_5_20_SNR",
    "EB_WSE_WEAK_20_40_NE", "EB_WSE_WEAK_20_40_SNR",
    "FB_FRE_HIGH_ENERGY_RATIO", "FB_FRE_HIGH_FREQUENCY_RANGE_HZ",
    "FB_FRE_HIGH_NE", "FB_FRE_HIGH_SNR",
    "FB_FRE_LOW_ENERGY_RATIO", "FB_FRE_LOW_FREQUENCY_RANGE_HZ",
    "FB_FRE_LOW_NE", "FB_FRE_LOW_SNR",
    "FB_FRE_MID_ENERGY_RATIO", "FB_FRE_MID_FREQUENCY_RANGE_HZ",
    "FB_FRE_MID_NE", "FB_FRE_MID_SNR",
    "FB_FRE_VERY_HIGH_ENERGY_RATIO", "FB_FRE_VERY_HIGH_FREQUENCY_RANGE_HZ",
    "FB_FRE_VERY_HIGH_NE", "FB_FRE_VERY_HIGH_SNR",
]

DASH = "—"

_params_cache: Dict[str, Optional[float]] = {}


def _resolve(path: str, repo_root: Path = _REPO_ROOT) -> Path:
    p = Path(path)
    return p if p.is_absolute() else (repo_root / p).resolve()


def _strip_seed(name: str) -> str:
    return re.sub(r"_seed[0-9]+_", "_seedN_", name)


def _load_metrics(collect_root: Path, name: str) -> Optional[Dict[str, Any]]:
    p = collect_root / name / "inference" / "metrics_summary.json"
    if not p.is_file():
        return None
    try:
        return json.loads(p.read_text())
    except json.JSONDecodeError:
        return None


def _count_params_m(ckpt_path: Path) -> Optional[float]:
    import torch

    ckpt = None
    for weights_only in (True, False):
        try:
            ckpt = torch.load(ckpt_path, map_location="cpu", weights_only=weights_only)
            break
        except Exception:
            continue
    if ckpt is None or not isinstance(ckpt, dict):
        return None
    sd = ckpt.get("model")
    if not isinstance(sd, dict):
        sd = ckpt.get("state_dict")
    if not isinstance(sd, dict):
        return None
    n = sum(int(t.numel()) for t in sd.values() if hasattr(t, "numel"))
    return n / 1e6


def _params_for_model(model_type: str, ckpt_candidates: List[Path]) -> Optional[float]:
    if model_type in _params_cache:
        return _params_cache[model_type]
    n = None
    for pt in ckpt_candidates:
        if pt.is_file():
            n = _count_params_m(pt)
            break
    _params_cache[model_type] = n
    return n


def _method_label(config_name: str, model_type: str) -> str:
    m = _NAME_RE.match(config_name)
    if m is None:
        return config_name
    rest = m.group(3)
    mask_mode, _, miss = rest.partition("_miss")
    if miss:
        return f"{model_type} ({mask_mode} {miss})"
    return f"{model_type} ({rest})"


def _fmt(mean: float, std: float, metric: str) -> str:
    if metric in ("mae", "mse", "rmse"):
        return f"{mean:.6f}±{std:.6f}"
    if metric in ("snr", "psnr", "ssim"):
        return f"{mean:.4f}±{std:.4f}"
    return f"{mean:.2f}±{std:.2f}"


def _column_value(per_seed: List[Dict[str, Any]], col: str) -> str:
    if col.endswith("FREQUENCY_RANGE_HZ"):
        return DASH
    key = col.lower()
    vals: List[float] = []
    for m in per_seed:
        if key == "rmse":
            mse = m.get("mse")
            v = math.sqrt(float(mse)) if mse is not None else None
        else:
            v = m.get(key)
        if v is None:
            continue
        try:
            vals.append(float(v))
        except (TypeError, ValueError):
            continue
    if not vals:
        return DASH
    mean = float(np.mean(vals))
    std = float(np.std(vals, ddof=1)) if len(vals) >= 2 else 0.0
    return _fmt(mean, std, key)


def _read_headers(wb) -> List[str]:
    if "Multiples" in wb.sheetnames:
        ws = wb["Multiples"]
        hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if hdr and str(hdr[0]).strip() == "Method":
            return [str(h) if h is not None else "" for h in hdr]
    return list(HEADERS)


def _build_sheet(wb, headers: List[str], rows: List[List[Any]]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if "Interpolation" in wb.sheetnames:
        del wb["Interpolation"]
    ws = wb.create_sheet("Interpolation")

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = left_align if ci == 1 else center_align

    ws.freeze_panes = "C2"
    from openpyxl.utils import get_column_letter
    for ci in range(1, len(headers) + 1):
        max_w = max(len(str(ws.cell(row=r, column=ci).value))
                    for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(ci)].width = max_w + 4


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config-dir", type=str, default="collected/configs")
    parser.add_argument("--collect-root", type=str, default="collected")
    parser.add_argument("--params-dir", type=str, default="collected/params")
    parser.add_argument("--output", type=str, default="batch_evaluation_part.xlsx")
    parser.add_argument(
        "--seeds", type=str, default="42 43 44",
        help="Space-separated seeds to aggregate (default '42 43 44').",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config_dir = _resolve(args.config_dir)
    collect_root = _resolve(args.collect_root)
    params_dir = _resolve(args.params_dir)
    output = _resolve(args.output)
    seeds = [int(s) for s in args.seeds.split()]

    if not config_dir.is_dir():
        print(f"ERROR: config dir not found: {config_dir}", file=sys.stderr)
        sys.exit(1)

    import openpyxl
    if output.is_file():
        wb = openpyxl.load_workbook(str(output))
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    headers = _read_headers(wb)

    configs = sorted(config_dir.glob("*.yaml"))
    if not configs:
        print(f"No configs found under {config_dir}.", file=sys.stderr)
        sys.exit(1)

    # Group configs by setting (seed stripped).
    settings: Dict[str, List[Path]] = {}
    for cfg in configs:
        settings.setdefault(_strip_seed(cfg.stem), []).append(cfg)

    rows: List[List[Any]] = []
    missing: List[str] = []
    for setting_key in sorted(settings):
        cfg_paths = settings[setting_key]
        first_cfg = cfg_paths[0]
        base_name = first_cfg.stem

        model_type = ""
        try:
            model_type = str(yaml.safe_load(first_cfg.read_text())["model"]["type"])
        except (OSError, KeyError, TypeError, yaml.YAMLError):
            pass

        label = _method_label(base_name, model_type)

        # Collect per-seed metrics.
        per_seed: List[Dict[str, Any]] = []
        for cfg in cfg_paths:
            name = cfg.stem
            m = _load_metrics(collect_root, name)
            if m is not None:
                per_seed.append(m)
        if not per_seed:
            missing.append(label)
            continue

        # Parameters (M): one cached load per model type from the first
        # available checkpoint.
        ckpt_candidates = [params_dir / f"{cfg.stem}.pt" for cfg in cfg_paths]
        params_m = _params_for_model(model_type, ckpt_candidates)

        row: List[Any] = [label, round(params_m, 2) if params_m is not None else DASH]
        for col in headers[2:]:
            row.append(_column_value(per_seed, col))
        rows.append(row)

    _build_sheet(wb, headers, rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print(f"Settings with results : {len(rows)}")
    print(f"Settings missing      : {len(missing)}")
    for lab in missing:
        print(f"  {lab}")
    print(f"Sheet 'Interpolation' written to: {output}")
    print("Model parameter counts (from checkpoints):")
    for m_key in sorted(_params_cache):
        n_m = _params_cache[m_key]
        print(f"  {m_key:<22s} {n_m:.2f}M" if n_m is not None else f"  {m_key}  n/a")


if __name__ == "__main__":
    main()
